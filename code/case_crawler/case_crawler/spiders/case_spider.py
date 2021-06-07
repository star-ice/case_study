import scrapy
import openpyxl as opxl
from bs4 import BeautifulSoup
from scrapy.http import Request
from scrapy.http import FormRequest
import json
import os
import case_crawler.settings as settings
from scrapy.crawler import CrawlerProcess
from scrapy.utils.project import get_project_settings
import time
import datetime
import requests

class CaseSpider(scrapy.Spider):
    name = "CaseSpider"
    allowed_domains = ["lawsdata.com"]
    pre_url = "https://www.lawsdata.com/api/compass-front/detail/getInstrumentById/"
    start_urls = "https://www.lawsdata.com/api/compass-auth/login/account"
    count = 0
    headers = settings.DEFAULT_REQUEST_HEADERS
    pre_time = datetime.datetime.now()
    aft_time = datetime.datetime.now()

    def __init__(self, dir):
        self.pre_dir = dir[0]
        self.dir_name = dir[1]
        self.dir_sname = dir[2]

    def start_urls_getter(self, file_name):
        start_urls = []
        case_ids = []
        # workbook = opxl.load_workbook("/Users/starice/Desktop/test.xlsx")
        print(file_name)
        workbook = opxl.load_workbook(file_name)
        sheet = workbook['案例']
        cells = sheet.iter_rows(max_col=2, max_row=sheet.max_row, min_col=2, min_row=sheet.min_row)
        for row in cells:
            for c in row:
                temp = c.value
                temp_str = str(temp)
                if "HYPERLINK" in temp_str:
                    start_url = temp_str[11:-1].split(",")[0][1:-1]
                    start_urls.append(start_url)
                    index_id = start_url.find("id=")
                    index_type = start_url.find("&type")
                    # print(index_id, index_type)
                    id = start_url[index_id+3:index_type]
                    case_ids.append(id)
        return start_urls, case_ids

    def start_requests(self, start_case=1):
        # start_urls = "https://www.lawsdata.com/api/compass-auth/login/account"
        payload = {"account": "13726286963", "password": "5598EF126B92325396B4F9062236F1AE"}
        res = requests.post(self.start_urls, data=json.dumps(payload))
        res_json = eval(res.text)
        self.headers['authentication'] = res_json['returnData']
        # yield FormRequest(self.start_urls, self.parse, method="POST", body=json.dumps(payload))
        base_url = "/Users/starice/OwnFiles/cityu/RA/"
        url = base_url + self.pre_dir + "/" + self.dir_name + "/" + self.dir_sname
        print(url)
        if not os.listdir(url):
            print(url, '目录为空!')
        else:
            files = os.listdir(url)
            for file in files:
                if file[-4:]=="xlsx":
                    print("file: ", file)
                    start_urls, case_ids = self.start_urls_getter(url + "/" + file)
                    for i in range(len(case_ids))[start_case-1:]:
                        yield Request(self.pre_url + case_ids[i], self.parse, meta={"id":case_ids[i], "url":url, \
                                                                                    "start_case":i}, headers=self.headers)

    def parse(self, response):
        result = response.text
        print(result)
        id = response.meta['id']
        url = response.meta['url']
        start_case = response.meta['start_case']
        # self.pre_time = response.meta['pre_time']
        self.aft_time = datetime.datetime.now()
        # print("aft_time: ", self.aft_time)
        if (self.aft_time - self.pre_time).seconds > 1700:
            print("在重新登录中！！！！！！！！请等待10秒！！！！！！")
            res = requests.post(self.start_urls, data=json.dumps({"account": "13726286963",
                                                       "password": "5598EF126B92325396B4F9062236F1AE"}))
            res_json = eval(res.text)
            # print(res_json)
            self.headers['authentication'] = res_json['returnData']
            time.sleep(10)
            print("重新登录后有了新的请求头！", self.headers)
            self.pre_time = datetime.datetime.now()
            # print("The renewed pre_time is: ", self.pre_time)
            self.aft_time = datetime.datetime.now()
            self.start_requests(start_case)
        #判断是否被封号
        if len(result) < 80:
            print("凉凉！要换号了！")
            print("已经下载了 ", self.count, " 个案例！")
            process.stop()

        file_name = url + "/json/" + id + ".json"
        # 将文件夹路径分割出来
        file_dir = os.path.split(file_name)[0]
        print(file_name, file_dir)

        # 判断文件夹路径是否存在，如果不存在，则创建，此处是创建多级目录
        if not os.path.isdir(file_dir):
            os.makedirs(file_dir)

        # 判断文件是否存在，不存在则创建
        # if not os.path.exists(file_name):
        # w是覆盖，a是追加，a+,w+是如不存在就创建
        with open(url + "/" + "json/" + id + ".json", 'w', encoding='utf-8') as file:
            file.write(json.dumps(result, indent=2, ensure_ascii=False))
        self.count += 1
        # print("写入成功！")


process = CrawlerProcess(get_project_settings())

pre_dir = ['type1', 'type2', 'type3', 'type4']
dir_name = ['2014', '2015', '2016', '2017', '2018', '2019', '2020']
dir_srname = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']
for i in pre_dir[1:2]:
    for j in dir_name[5:6]:
        for k in dir_srname[3:4]:
            # time.sleep(5)
            process.crawl('CaseSpider', dir = [i, j, k])
process.start()  # the script will block here until the crawling is finished
process.stop()
